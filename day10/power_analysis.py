import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
import numpy as np
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

def create_power_analysis_excel():
    # CSV 파일 읽기 (인코딩 문제 해결)
    try:
        # UTF-8로 시도
        df = pd.read_csv('한국수력원자력(주)_전원별 전력판매량 현황_20250430.csv', encoding='utf-8')
    except:
        try:
            # CP949로 시도
            df = pd.read_csv('한국수력원자력(주)_전원별 전력판매량 현황_20250430.csv', encoding='cp949')
        except:
            # EUC-KR로 시도
            df = pd.read_csv('한국수력원자력(주)_전원별 전력판매량 현황_20250430.csv', encoding='euc-kr')
    
    # 컬럼명 설정 (한글이 깨진 경우를 대비)
    df.columns = ['날짜', '원자력', '수력', '양수', '신재생', '합계']
    
    # 날짜 컬럼을 datetime으로 변환
    df['날짜'] = pd.to_datetime(df['날짜'])
    
    # 년도와 월 컬럼 추가
    df['년도'] = df['날짜'].dt.year
    df['월'] = df['날짜'].dt.month
    
    # 엑셀 파일 생성
    with pd.ExcelWriter('전력판매량_분석_대시보드.xlsx', engine='openpyxl') as writer:
        
        # 1. 원본 데이터 시트
        df.to_excel(writer, sheet_name='원본데이터', index=False)
        
        # 2. 년도별 분석 시트
        yearly_analysis = df.groupby('년도').agg({
            '원자력': 'sum',
            '수력': 'sum', 
            '양수': 'sum',
            '신재생': 'sum',
            '합계': 'sum'
        }).round(2)
        
        yearly_analysis.to_excel(writer, sheet_name='년도별분석')
        
        # 3. 월별 분석 시트
        monthly_analysis = df.groupby(['년도', '월']).agg({
            '원자력': 'sum',
            '수력': 'sum',
            '양수': 'sum', 
            '신재생': 'sum',
            '합계': 'sum'
        }).round(2)
        
        monthly_analysis.to_excel(writer, sheet_name='월별분석')
        
        # 4. 월별 평균 분석 시트
        monthly_avg = df.groupby('월').agg({
            '원자력': 'mean',
            '수력': 'mean',
            '양수': 'mean',
            '신재생': 'mean',
            '합계': 'mean'
        }).round(2)
        
        monthly_avg.to_excel(writer, sheet_name='월별평균')
        
        # 5. 전원별 비율 분석 시트
        df_ratio = df.copy()
        df_ratio['원자력_비율'] = (df_ratio['원자력'] / df_ratio['합계'] * 100).round(2)
        df_ratio['수력_비율'] = (df_ratio['수력'] / df_ratio['합계'] * 100).round(2)
        df_ratio['양수_비율'] = (df_ratio['양수'] / df_ratio['합계'] * 100).round(2)
        df_ratio['신재생_비율'] = (df_ratio['신재생'] / df_ratio['합계'] * 100).round(2)
        
        ratio_analysis = df_ratio[['날짜', '원자력_비율', '수력_비율', '양수_비율', '신재생_비율']]
        ratio_analysis.to_excel(writer, sheet_name='비율분석', index=False)
        
        # 6. 요약 통계 시트
        summary_stats = df[['원자력', '수력', '양수', '신재생', '합계']].describe().round(2)
        summary_stats.to_excel(writer, sheet_name='요약통계')
    
    # 차트 생성
    create_charts(df)
    
    print("전력판매량 분석 엑셀 파일이 생성되었습니다: 전력판매량_분석_대시보드.xlsx")

def create_charts(df):
    """다양한 차트를 생성하고 저장"""
    
    # 1. 년도별 전원별 판매량 추이
    plt.figure(figsize=(15, 8))
    yearly_data = df.groupby('년도')[['원자력', '수력', '양수', '신재생']].sum()
    
    for column in ['원자력', '수력', '양수', '신재생']:
        plt.plot(yearly_data.index, yearly_data[column], marker='o', linewidth=2, label=column)
    
    plt.title('년도별 전원별 전력판매량 추이', fontsize=16, fontweight='bold')
    plt.xlabel('년도', fontsize=12)
    plt.ylabel('판매량 (GWh)', fontsize=12)
    plt.legend(fontsize=10)
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('년도별_전원별_판매량_추이.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    # 2. 월별 평균 판매량
    plt.figure(figsize=(15, 8))
    monthly_avg = df.groupby('월')[['원자력', '수력', '양수', '신재생']].mean()
    
    x = np.arange(len(monthly_avg))
    width = 0.2
    
    plt.bar(x - width*1.5, monthly_avg['원자력'], width, label='원자력', alpha=0.8)
    plt.bar(x - width*0.5, monthly_avg['수력'], width, label='수력', alpha=0.8)
    plt.bar(x + width*0.5, monthly_avg['양수'], width, label='양수', alpha=0.8)
    plt.bar(x + width*1.5, monthly_avg['신재생'], width, label='신재생', alpha=0.8)
    
    plt.title('월별 평균 전원별 판매량', fontsize=16, fontweight='bold')
    plt.xlabel('월', fontsize=12)
    plt.ylabel('평균 판매량 (GWh)', fontsize=12)
    plt.legend(fontsize=10)
    plt.xticks(x, [f'{i}월' for i in monthly_avg.index])
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('월별_평균_판매량.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    # 3. 전원별 비율 변화 (스택 영역 차트)
    plt.figure(figsize=(15, 8))
    yearly_ratio = df.groupby('년도').apply(lambda x: pd.Series({
        '원자력_비율': (x['원자력'].sum() / x['합계'].sum() * 100),
        '수력_비율': (x['수력'].sum() / x['합계'].sum() * 100),
        '양수_비율': (x['양수'].sum() / x['합계'].sum() * 100),
        '신재생_비율': (x['신재생'].sum() / x['합계'].sum() * 100)
    }))
    
    plt.stackplot(yearly_ratio.index, 
                  yearly_ratio['원자력_비율'], 
                  yearly_ratio['수력_비율'],
                  yearly_ratio['양수_비율'],
                  yearly_ratio['신재생_비율'],
                  labels=['원자력', '수력', '양수', '신재생'],
                  alpha=0.8)
    
    plt.title('년도별 전원별 비율 변화', fontsize=16, fontweight='bold')
    plt.xlabel('년도', fontsize=12)
    plt.ylabel('비율 (%)', fontsize=12)
    plt.legend(fontsize=10)
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('년도별_전원별_비율_변화.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    # 4. 월별 판매량 히트맵
    plt.figure(figsize=(12, 8))
    monthly_pivot = df.pivot_table(values='합계', index='월', columns='년도', aggfunc='sum')
    
    sns.heatmap(monthly_pivot, annot=True, fmt='.0f', cmap='YlOrRd', 
                cbar_kws={'label': '판매량 (GWh)'})
    plt.title('월별/년도별 총 판매량 히트맵', fontsize=16, fontweight='bold')
    plt.xlabel('년도', fontsize=12)
    plt.ylabel('월', fontsize=12)
    plt.tight_layout()
    plt.savefig('월별_년도별_히트맵.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    # 5. 신재생에너지 성장 추이
    plt.figure(figsize=(15, 8))
    renewable_growth = df.groupby('년도')['신재생'].sum()
    
    plt.plot(renewable_growth.index, renewable_growth.values, marker='o', 
             linewidth=3, markersize=8, color='green')
    plt.fill_between(renewable_growth.index, renewable_growth.values, alpha=0.3, color='green')
    
    plt.title('신재생에너지 판매량 성장 추이', fontsize=16, fontweight='bold')
    plt.xlabel('년도', fontsize=12)
    plt.ylabel('신재생에너지 판매량 (GWh)', fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('신재생에너지_성장_추이.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print("차트 이미지들이 생성되었습니다.")

def add_charts_to_excel():
    """엑셀 파일에 차트를 추가하는 함수"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    
    # 엑셀 파일 로드
    wb = load_workbook('전력판매량_분석_대시보드.xlsx')
    
    # 대시보드 시트 생성
    if '대시보드' in wb.sheetnames:
        ws = wb['대시보드']
    else:
        ws = wb.create_sheet('대시보드')
    
    # 제목 추가
    ws['A1'] = '한국수력원자력(주) 전원별 전력판매량 분석 대시보드'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    
    # 차트 이미지 추가
    chart_images = [
        ('년도별_전원별_판매량_추이.png', 'A3'),
        ('월별_평균_판매량.png', 'A25'),
        ('년도별_전원별_비율_변화.png', 'A47'),
        ('월별_년도별_히트맵.png', 'A69'),
        ('신재생에너지_성장_추이.png', 'A91')
    ]
    
    for img_file, cell in chart_images:
        try:
            img = Image(img_file)
            img.width = 600
            img.height = 400
            ws.add_image(img, cell)
        except:
            print(f"이미지 {img_file} 추가 실패")
    
    # 파일 저장
    wb.save('전력판매량_분석_대시보드.xlsx')
    print("엑셀 파일에 차트가 추가되었습니다.")

if __name__ == "__main__":
    # 메인 실행
    create_power_analysis_excel()
    add_charts_to_excel()
    print("분석이 완료되었습니다!") 