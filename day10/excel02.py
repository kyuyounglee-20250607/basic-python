from datetime import datetime, timedelta
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

# 샘플 데이터 생성
dates = [(datetime.now()-timedelta(days=i)).strftime("%Y-%m-%d") 
                for i in range(30,0,-1)]
data = {
        '날짜' : dates,
        '매출':[ 1000+ i*50 +(i%7)*100 for i in range(30)],
        '방문자수':[ 100+ i*5 +(i%3)*10 for i in range(30)],
        '전환율':[ round(2.5 + (i%5)*0.5,2) for i in range(30)],
        '지역':[['서울','부산','대구','인천','광주'][i%5] for i in range(30)]
    }

# DataFrame 생성
df = pd.DataFrame(data)
df['날짜'] = pd.to_datetime(df['날짜'])
df['년도'] = df['날짜'].dt.year
df['월'] = df['날짜'].dt.month

excel_filename = '매출_분석_대시보드.xlsx'

# 엑셀 파일 생성 및 데이터 시트 저장
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='데이터', index=False)

# 차트 생성 함수
def save_charts(group_df, group_name, value_col, chart_prefix):
    chart_files = []
    # 막대 차트
    plt.figure(figsize=(7,4))
    group_df.plot(kind='bar', x=group_name, y=value_col, legend=False, color='skyblue')
    plt.title(f'{group_name}별 {value_col} (Bar)')
    plt.xlabel(group_name)
    plt.ylabel(value_col)
    plt.tight_layout()
    bar_path = f'{chart_prefix}_bar.png'
    plt.savefig(bar_path, dpi=200)
    plt.close()
    chart_files.append(bar_path)
    # 라인 차트
    plt.figure(figsize=(7,4))
    group_df.plot(kind='line', x=group_name, y=value_col, legend=False, marker='o', color='green')
    plt.title(f'{group_name}별 {value_col} (Line)')
    plt.xlabel(group_name)
    plt.ylabel(value_col)
    plt.tight_layout()
    line_path = f'{chart_prefix}_line.png'
    plt.savefig(line_path, dpi=200)
    plt.close()
    chart_files.append(line_path)
    # 파이 차트
    plt.figure(figsize=(5,5))
    plt.pie(group_df[value_col], labels=group_df[group_name], autopct='%1.1f%%', startangle=90)
    plt.title(f'{group_name}별 {value_col} (Pie)')
    plt.tight_layout()
    pie_path = f'{chart_prefix}_pie.png'
    plt.savefig(pie_path, dpi=200)
    plt.close()
    chart_files.append(pie_path)
    return chart_files

# 집계 및 차트 생성/엑셀 삽입 함수
def make_sheet_and_charts(wb, groupby_col, value_col, sheet_name):
    # 집계
    group_df = df.groupby(groupby_col)[value_col].sum().reset_index()
    # 시트 생성
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    # 데이터 기록
    for i, col in enumerate(group_df.columns):
        ws.cell(row=1, column=i+1, value=col)
    for row_idx, row in enumerate(group_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    # 차트 생성 및 삽입
    chart_files = save_charts(group_df, groupby_col, value_col, f'{sheet_name}_{value_col}')
    anchors = ['E2', 'E20', 'M2']
    for img_path, anchor in zip(chart_files, anchors):
        img = XLImage(img_path)
        img.width = img.width // 2
        img.height = img.height // 2
        img.anchor = anchor
        ws.add_image(img)
    # 임시 이미지 삭제는 마지막에 일괄 처리
    return chart_files

# 엑셀 파일 열기
wb = load_workbook(excel_filename)

# 지역별, 년도별, 월별 시트 및 차트 생성
all_temp_imgs = []
for value_col in ['매출', '방문자수', '전환율']:
    all_temp_imgs += make_sheet_and_charts(wb, '지역', value_col, f'지역별_{value_col}')
    all_temp_imgs += make_sheet_and_charts(wb, '년도', value_col, f'년도별_{value_col}')
    all_temp_imgs += make_sheet_and_charts(wb, '월', value_col, f'월별_{value_col}')

wb.save(excel_filename)

# 임시 이미지 파일 삭제
for img_path in all_temp_imgs:
    if os.path.exists(img_path):
        os.remove(img_path)

print(f"엑셀 파일 '{excel_filename}'에 지역별/년도별/월별 시트와 차트가 모두 생성되었습니다.")                